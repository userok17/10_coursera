import requests
from bs4 import BeautifulSoup
import openpyxl
import random
from datetime import datetime
import os
import webbrowser

os.chdir(os.path.dirname(os.path.abspath(__file__)))

        
def get_courses_list(count_courses):
    url = 'https://www.coursera.org/sitemap~www~courses.xml'
    request = requests.get(url)
    soup = BeautifulSoup(request.text, 'lxml')
    tags_loc = random.sample(soup.find_all('loc'), count_courses)
    courses_list = []
    for url in tags_loc:
        course_info = get_course_info(url.text)
        courses_list.append(course_info)
    return courses_list   

def get_course_info(url):
    request = requests.get(url)
    soup = BeautifulSoup(request.text, 'lxml')
    title = soup.find('h1', class_='title').text
    language = soup.find('div', class_='language-info').text
    start_date = soup.find('div', class_='startdate').text
    weeks = len(soup.select('div.week'))
    element_raiting = soup.find('div', class_='ratings-text')
    if element_raiting:
        raiting = element_raiting.text
    else:
        raiting = ''
    course_info = {
        'url': url,
        'title': title,
        'language': language,
        'start_date': start_date,
        'weeks': weeks,
        'raiting': raiting
    }
    return course_info

def output_courses_info_to_xlsx(filepath, courses_list):
        width_columns = 25
        offset_row = 2
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1, value='Title')
        sheet.cell(row=1, column=2, value='Language')
        sheet.cell(row=1, column=3, value='Start date')
        sheet.cell(row=1, column=4, value='Weeks')
        sheet.cell(row=1, column=5, value='Raiting')
        sheet.cell(row=1, column=6, value='Url')

        
        for idx, course_info in enumerate(courses_list):
            row = idx + offset_row
            sheet.cell(row=row, column=1, value=course_info['title'])
            sheet.cell(row=row, column=2, value=course_info['language'])
            sheet.cell(row=row, column=3, value=course_info['start_date'])
            sheet.cell(row=row, column=4, value=course_info['weeks'])
            sheet.cell(row=row, column=5, value=course_info['raiting'])
            sheet.cell(row=row, column=6, value=course_info['url'])
            

        sheet.column_dimensions['A'].width = width_columns
        sheet.column_dimensions['B'].width = width_columns
        sheet.column_dimensions['C'].width = width_columns
        sheet.column_dimensions['D'].width = width_columns
        sheet.column_dimensions['E'].width = width_columns
        sheet.column_dimensions['F'].width = width_columns
        
        workbook.save(filepath)

def create_folder(folder):
    if not os.path.isdir(folder):
        os.mkdir(folder)

def generate_filename():
    now = datetime.now()
    date_string = now.strftime('%d-%m-%Y-%H-%M-%S')
    filename = '{}.xlsx'.format(date_string)
    return filename

def open_xlsx(filepath):
    webbrowser.open(filepath)

def main():
    print('Сканирование курсов')
    
    count_courses = 20
    courses_list = get_courses_list(count_courses)

    folder = 'data'
    create_folder(folder)
    
    filename = generate_filename()
    filepath = os.path.join(folder, filename)

    output_courses_info_to_xlsx(filepath, courses_list)

    open_xlsx(filepath)
    
    print('Сканирование курсов завершено')
    
if __name__ == '__main__':
    main()
    
    
    
